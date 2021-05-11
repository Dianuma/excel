try:
    import Tkinter as tk
except:
    import tkinter as tk
import tkinter.messagebox
import tkinter.ttk as ttk

import sys
import os
import numpy as np
import time

import openpyxl
import pyexcel as p
import pyexcel_xls
import pyexcel_xlsx
import pyexcel_xlsxw
import math

#font=("UD Digi Kyokasho N-B", 20, "bold")
font=("TkDefaultFont",10)

class SampleApp(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.title("엑셀 변환기")
        self.geometry("750x530+100+100")
        self.resizable(False, False)
        self._frame = None
        self.temp=[]
        self.temp_2=[]
        self.temp_3={}
        self.temp_4=[]
        self.selec_temp={}
        self.ID_Number=0
        self.switch_frame(StartPage)

    def switch_frame(self, frame_class):
        new_frame = frame_class(self)
        if self._frame is not None:
            self._frame.destroy()
        self._frame = new_frame
        self._frame.pack()   

    def _exit(self):
        sys.exit()

class StartPage(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.master=master
        self.file_name=os.listdir('엑셀 넣는 곳')
        self.all_values = {}
        self.key=self.key_sort()
        tk.Button(self, text="EXIT", font=("TkDefaultFont",15,"bold"),command=lambda: self.master._exit()).pack(side="bottom",anchor="e")
        self.get_item()

    def get_item(self):
        tk.Label(self, text="\n변환할 파일을 모두 선택해주세요",font=("TkDefaultFont",13,"bold")).pack()
        frame_1=tk.Frame(self,width=600,height=200)
        frame_1.pack()
        scrollbar = tk.Scrollbar(frame_1,orient=tk.HORIZONTAL)
        scrollbar_2 = tk.Scrollbar(frame_1)
        text = tk.Text(frame_1,relief="flat",xscrollcommand=scrollbar.set,yscrollcommand=scrollbar_2.set,borderwidth=0)
        scrollbar.config(command=text.xview,)
        scrollbar.pack(side="bottom",fill="x")
        scrollbar_2.config(command=text.yview)
        scrollbar_2.pack(side="right",fill="y")
        text.pack(side="top",fill="both",expand=True)
        self.Var=[]
        self.check_box=[]
        for i in range(len(self.key)):
            self.Var.append(tk.IntVar())
            cb = tk.Checkbutton(frame_1,text=self.key[i],variable=self.Var[i], font=font,padx=0,pady=0,bd=0,bg="white",borderwidth =0)
            self.check_box.append(cb)
            text.window_create("end", window=cb)
            text.insert("end", "\n") 
        scrollbar["command"]=text.xview
        scrollbar_2["command"]=text.yview
        tk.Button(self,text='모두 선택',command=self.set_all).pack()
        tk.Button(self,text='모두 선택 취소',command=self.deselect_all).pack()
        tk.Button(self,text='선택 완료',command=self.item).pack()

    def key_sort(self):
        key_int=[]
        key_str=[]
        for i in self.file_name:
            if i.split('.')[1]=="xlsx" or i.split('.')[1]=="xls":
                try :
                    int(i.split('.')[0])
                    key_int.append(i)
                except :
                    key_str.append(i)
        key=sorted(key_int,key=lambda fname: int(fname.split('.')[0]))+sorted(key_str)
        return key

    def item(self):
        self.master.temp=[self.key[i] for i in range(len(self.Var)) if self.Var[i].get()==1]
        self.master.switch_frame(PageOne)
 
    def set_all(self):
        [i.select() for i in self.check_box]
        
    def deselect_all(self):
        [i.deselect() for i in self.check_box]
          
class PageOne(tk.Frame):
    def __init__(self, master):
        self.master=master
        tk.Frame.__init__(self, master)
        tk.Frame.configure(self)
        self.worksheet_CP=[]
        self.worksheet_NH=[]
        self.worksheet_SH=[]
        self.serial_number=[[7,5],[8,6],[2,4]]
        self.serial_count=0
        self.load_excel()
        tk.Label(self, text="\n상대 계좌 번호 항목이 존재하지 않거나\n상대 계좌 번호가 같은 건이 존재하지 않는 파일의 경우 자동으로 무시됩니다.",font=("TkDefaultFont",13,"bold"),fg="blue").pack(side="bottom")
        tk.Button(self, text="EXIT", font=("TkDefaultFont",15,"bold"),command=lambda: self.master._exit()).pack(side="bottom",anchor="e")
        
        temp_temp=list(self.master.temp_2)
        while len(temp_temp)>=1:
            self.pre=temp_temp.pop(0)
            self.all_values=self.master.temp_3[self.pre]
            self.data=np.array(self.all_values)[6:]
            self.selec=self.change_second()
            if len(self.selec.keys())>=1:
                self.master.selec_temp[self.pre]=self.selec
            else:
                self.master.temp_2.remove(self.pre)

        self.get_item()

    def change_second(self):
        set_list=list(set(self.data[:,6]))
        if self.data[len(self.data)-1,5]==None:
            length=len(self.data)-1
        elif self.data[len(self.data)-1,5]!=None:length=len(self.data)
        total=0
        case_by_total=0
        selec=[]
        for number in set_list:
            if number!=None:
                business_number=[]
                for count in range(length):
                    if self.data[count,5]!=None:
                        if self.data[count,6]==number:
                            business_number.append(count)
                        count+=1
                business_name=[]
                for i in business_number:
                    business_name.append(self.data[i,5])
                    set_business_name=list(set(business_name))
                if len(set_business_name)>=1.5:
                    case_by_total+=len(business_number)
                    selec.append([business_name,business_number])
                total+=len(business_number)
        selection={}
        for i in selec:
            temp_selec={}
            for j in range(len(i[0])):
                temp_selec[i[0][j]]=[]
            for j in range(len(i[0])):
                temp_selec[i[0][j]].append(i[1][j])
            temp_selec=self.dictionary_sort(temp_selec)
            selection[min(i[0],key=len)]=temp_selec
        selection=self.dictionary_sort(selection)
        
        return selection

    def dictionary_sort(self,dic):
        A=sorted(dic.keys(),key=len)
        B={}
        for i in A:
            B[i]=dic[i]
        return B

    def get_item(self):
        tk.Label(self, text="\n상대 계좌 번호를 이용해 변환할 파일을 모두 선택해주세요",font=("TkDefaultFont",13,"bold")).pack()
        frame_1=tk.Frame(self,width=600,height=200)
        frame_1.pack(side="left")
        scrollbar = tk.Scrollbar(frame_1,orient=tk.HORIZONTAL)
        scrollbar_2 = tk.Scrollbar(frame_1)
        text = tk.Text(frame_1,relief="flat",xscrollcommand=scrollbar.set,yscrollcommand=scrollbar_2.set,borderwidth=0)
        scrollbar.config(command=text.xview,)
        scrollbar.pack(side="bottom",fill="x")
        scrollbar_2.config(command=text.yview)
        scrollbar_2.pack(side="right",fill="y")
        text.pack(side="top",fill="both",expand=True)
        self.Var=[]
        self.check_box=[]
        for i in range(len(self.master.temp_2)):
            self.Var.append(tk.IntVar())
            cb = tk.Checkbutton(frame_1,text=self.master.temp_2[i],variable=self.Var[i], font=font,padx=0,pady=0,bd=0,bg="white",borderwidth =0)
            self.check_box.append(cb)
            text.window_create("end", window=cb)
            text.insert("end", "\n") 
        scrollbar["command"]=text.xview
        scrollbar_2["command"]=text.yview
        tk.Button(self,text='모두 선택',command=self.set_all).pack()
        tk.Button(self,text='모두 선택 취소',command=self.deselect_all).pack()
        tk.Button(self,text='선택 완료',command=self.item).pack()

    def item(self):
        self.master.temp_4=[self.master.temp_2[i] for i in range(len(self.Var)) if self.Var[i].get()==1]
        self.master.switch_frame(PageTwo)

    def set_all(self):
        [i.select() for i in self.check_box]
        
    def deselect_all(self):
        [i.deselect() for i in self.check_box]

    def load_excel(self):
        for pre in self.master.temp:
            if (pre).split(".")[1]=="xls":
                try:
                    p.save_book_as(file_name='엑셀 넣는 곳\\'+pre, dest_file_name=pre+'x')
                    pre_save=pre+'x'
                    workbook=openpyxl.load_workbook(pre+'x')
                    worksheed=workbook[workbook.sheetnames[0]]
                    os.remove(pre+'x')
                except:
                    tk.messagebox.showerror("오류","엑셀파일이 제대로 된 파일인지 확인해 주세요.")
            elif (pre).split(".")[1]=="xlsx":
                try:
                    workbook=openpyxl.load_workbook('엑셀 넣는 곳\\'+pre)
                    pre_save=pre
                    worksheed=workbook[workbook.sheetnames[0]]
                except:
                    tk.messagebox.showerror("오류","엑셀파일이 제대로 된 파일인지 확인해 주세요.")
            
            all_values = []
            serial_number=[[7,5],[8,6],[2,4]]
            for row in worksheed.rows:
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)

            if all_values[5][5]=="거래내용":
                self.master.temp_2.append(pre)
                self.serial_count=0
            elif all_values[6][6]=="거래기록사항":
                self.serial_count=1
            elif all_values[0][4]=="내용":
                self.serial_count=2

            count=self.serial_number[self.serial_count][0]
            after_cell="%s%d"%(chr(ord("A")+len(all_values[self.serial_number[self.serial_count][0]-2])),count)
            before_cell="%s%d"%(chr(ord("A")+self.serial_number[self.serial_count][1]),count)
            for i in range(len(all_values)-self.serial_number[self.serial_count][0]+1):
                worksheed[after_cell].value=worksheed[before_cell].value
                count+=1
                after_cell="%s%d"%(chr(ord("A")+len(all_values[self.serial_number[self.serial_count][0]-2])),count)
                before_cell="%s%d"%(chr(ord("A")+self.serial_number[self.serial_count][1]),count)

            all_values = []
            for row in worksheed.rows:
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                all_values.append(row_value)
            data=np.array(all_values)[self.serial_number[self.serial_count][0]-1:]

            #필요 없는 문자열 제거

            num=['１','２','３','４','５','６','７','８','９','０','1','2','3','4','5','6','7','8','9','0']
            mon=['상','하','월']

            delete_file=open("삭제 단어 목록.txt", encoding='UTF8')
            delete_=delete_file.read()
            dele=[]
            if delete_:
                dele=list((map(str,delete_.split("\n"))))

            change_file=open("변환 단어 목록.txt", encoding='UTF8')
            change_=change_file.read()
            chan=[]
            if change_:
                chan=[[i.split("//")[0],i.split("//")[1]] for i in list((map(str,change_.split("\n"))))]
            
            replace_file=open("수정 단어 목록.txt", encoding='UTF8')
            replace_=replace_file.read()
            repl=[]
            if replace_:
                repl=[[i.split("//")[0],i.split("//")[1]] for i in list((map(str,replace_.split("\n"))))]
            
            judge_file=open("기본 적용.txt", encoding='UTF8')
            judge=[str(i.split("=")[1]).replace(" ","") for i in list((map(str,judge_file.read().split("\n"))))]
            deli=[]
            if all_values[self.serial_number[self.serial_count][0]-2][0]=="거래일시":
                YY=data[0,0][:4]
            elif all_values[self.serial_number[self.serial_count][0]-2][1]=="거래일시":
                YY=data[0,1][:4]

            for i in num:
                for j in mon:
                    deli.append(i+j)

            count=0
            for i in data[:,self.serial_number[self.serial_count][1]]:
                if i!=None:
                    if judge[1]=='1':
                        for j in deli:
                            i=i.replace(j,'')
                    if judge[0]=='1':
                        for j in num:
                            i=i.replace(j,'')
                    for j in dele:
                        i=i.replace(j,'')
                    if judge[2]=='1':
                        if ("년결산" or "년 결산") in i:
                            i=YY+"년결산"
                    if judge[3]=='1':
                        for j in ['（주）','주）','（주','(주)','㈜','주식회사)','주식)','주)','(주식회사','(주식','(주','주식회사','주식회','주식']:
                            if j in i:
                                i="㈜"+i.replace(j,'')
                        if "㈜" in i:
                            i=i.replace("㈜","(주)")
                    for j in range(len(chan)):
                        if chan[j][0] in i:
                            i=chan[j][1]
                    for j in range(len(repl)):
                        i=i.replace(repl[j][0],repl[j][1])
                    if "()" in i:
                        i=i.replace("()","")
                data[count,self.serial_number[self.serial_count][1]]=i
                count+=1

            for i in range(len(data[:,self.serial_number[self.serial_count][1]])):
                if data[i,self.serial_number[self.serial_count][1]] != None:
                    k=self.serial_number[self.serial_count][0]+i-1
                    all_values[k][self.serial_number[self.serial_count][1]]=data[i,self.serial_number[self.serial_count][1]]

            if self.serial_count==0:
                self.master.temp_3[pre]=all_values


            wb=openpyxl.Workbook()
            ws=wb.active
            for i in all_values:
                ws.append(i)
            
            wb.save('엑셀 나오는 곳\\수정 후_'+pre_save)
class PageTwo(tk.Frame):
    def __init__(self, master):
        self.master=master
        tk.Frame.__init__(self, master)
        tk.Frame.configure(self)

        self.pre=self.master.temp_4.pop(0)
        self.all_values=self.master.temp_3[self.pre]
        self.data=np.array(self.all_values)[6:]
        self.selec=self.master.selec_temp[self.pre]
        tk.Label(self,text=self.pre,font=("TkDefaultFont",15,"bold")).pack()
        self.top_frame=tk.Frame(self, relief="sunken", bd=2)
        self.top_frame.pack(side="top",fill="both",expand=True)
        self.bottom_frame=tk.Frame(self, relief="sunken", bd=2)
        self.bottom_frame.pack(side="bottom",fill="both",expand=True)
        self.frame1=None
        self.frame2=None
        self.frame3=None
        self.frame4=None
        self.jud_frame2=None
        self.jud_frame3=None
        self.listbox1=None
        self.listbox2=None

        self.Frame1()
        self.Frame2()
        self.Frame3()
        self.Frame4()

    def delete_Frame(self):
        try :
            self.frame1.destroy()
            self.frame2.destroy()
            self.frame3.destroy()
            self.frame4.destroy()
            self.destroy()
        except :
            k=None

    def Frame1(self):   
        try :
            self.frame1.destroy()
            self.frame1=tk.Frame(self.top_frame, relief="sunken", bd=2, bg='white')
            self.frame1.pack(side="left",fill="both",expand=True)
        except :
            self.frame1=tk.Frame(self.top_frame, relief="sunken", bd=2, bg='white')
            self.frame1.pack(side="left",fill="both",expand=True)

        scrollbar=tk.Scrollbar(self.frame1)
        scrollbar.pack(side="right",fill="y")
        scrollbar_2=tk.Scrollbar(self.frame1,orient=tk.HORIZONTAL)
        scrollbar_2.pack(side="bottom",fill="x")
        self.listbox1=tk.Listbox(self.frame1, width=25,height=18, selectmode="extended", xscrollcommand=scrollbar_2.set,yscrollcommand=scrollbar.set,font=("TkDefaultFont",10))
        for i in range(len(self.selec.keys())):
            self.listbox1.insert(i,list(self.selec.keys())[i])
        self.listbox1.bind('<Double-1>',self.Frame1_clickevent)
        self.listbox1.pack(fill="both",expand=True)
        scrollbar["command"]=self.listbox1.yview
        scrollbar_2["command"]=self.listbox1.xview

    def Frame1_clickevent(self,event):
        self.jud_frame2=str(self.listbox1.selection_get())
        self.jud_frame3=None
        self.Frame2()
        self.Frame3()

    def Frame2(self):
        try :
            self.frame2.destroy()
            self.frame2=tk.Frame(self.top_frame, relief="sunken", bd=2, bg='white')
            self.frame2.pack(side="right",fill="both",expand=True)
        except :
            self.frame2=tk.Frame(self.top_frame, relief="sunken", bd=2, bg='white')
            self.frame2.pack(side="right",fill="both",expand=True)     

        if self.jud_frame2!=None:
            scrollbar=tk.Scrollbar(self.frame2)
            scrollbar.pack(side="right",fill="y")
            scrollbar_2=tk.Scrollbar(self.frame2,orient=tk.HORIZONTAL)
            scrollbar_2.pack(side="bottom",fill="x")
            self.listbox2=tk.Listbox(self.frame2, width=75,height=18, selectmode="extended", xscrollcommand=scrollbar_2.set,yscrollcommand=scrollbar.set,font=("TkDefaultFont",10))
            for i in range(len(self.selec[self.jud_frame2].keys())):
                self.listbox2.insert(i,list(self.selec[self.jud_frame2].keys())[i])
            self.listbox2.bind('<Double-1>',self.Frame2_clickevent)
            self.listbox2.pack(fill="both",expand=True)
            scrollbar["command"]=self.listbox2.yview
            scrollbar_2["command"]=self.listbox2.xview
        elif self.jud_frame2==None:
            scrollbar=tk.Scrollbar(self.frame2)
            scrollbar.pack(side="right",fill="y")
            scrollbar_2=tk.Scrollbar(self.frame2,orient=tk.HORIZONTAL)
            scrollbar_2.pack(side="bottom",fill="x")
            self.listbox2=tk.Listbox(self.frame2, width=75,height=18, selectmode="extended", xscrollcommand=scrollbar_2.set,yscrollcommand=scrollbar.set,font=("TkDefaultFont",10))
            self.listbox2.pack(fill="both",expand=True)
            scrollbar["command"]=self.listbox2.yview
            scrollbar_2["command"]=self.listbox2.xview

    def Frame2_clickevent(self,event):
        self.jud_frame3=str(self.listbox2.selection_get())
        self.selec[self.jud_frame3]=self.selec.pop(self.jud_frame2)
        self.jud_frame2=self.jud_frame3
        self.Frame1()    

    def Frame3(self):
        try :
            self.frame3.destroy()
            self.frame3=tk.Frame(self.bottom_frame, width=650,height=200, relief="sunken", bd=2, bg='white')
            self.frame3.pack(side="left",fill="both",expand=True)
        except :
            self.frame3=tk.Frame(self.bottom_frame, width=650,height=200, relief="sunken", bd=2, bg='white')
            self.frame3.pack(side="left",fill="both",expand=True) 

        if self.jud_frame3!=None:
            column=[]
            column_name=[]
            for i in range(len(self.all_values[5])):
                column.append(i+1)
                column_name.append(self.all_values[5][i])
            if column_name[len(column_name)-1]==None:
                column_name[len(column_name)-1]="백업 내용"
            len_treelist=[]
            for i in range(len((self.selec[self.jud_frame2])[self.jud_frame3])):
                j=list(self.data[((self.selec[self.jud_frame2])[self.jud_frame3])[i]])
                for k in range(len(j)):
                    if len(len_treelist)<len(j):
                        len_treelist.append(str(j[k]))
                    else:
                        len_treelist[k]=str(max([len_treelist[k],str(j[k])],key=len))

            len_treeview=[]
            for i in len_treelist:
                lenght=0
                for j in str(i):
                    if (ord("a")<=ord(j) and ord("z")>=ord(j)) or (ord("0")<=ord(j) and ord("9")>=ord(j)) or ord(j)==45 or ord(j)==58:
                        lenght+=9
                    elif (ord("A")<=ord(j) and ord("Z")>=ord(j)):
                        lenght+=12
                    else : lenght+=17
                len_treeview.append(lenght)
            
            column_name_len_treeview=[]
            for i in column_name:
                lenght=0
                for j in str(i):
                    if (ord("a")<=ord(j) and ord("z")>=ord(j)) or (ord("0")<=ord(j) and ord("9")>=ord(j)) or ord(j)==45 or ord(j)==58:
                        lenght+=9
                    elif (ord("A")<=ord(j) and ord("Z")>=ord(j)):
                        lenght+=12
                    else : lenght+=17
                column_name_len_treeview.append(lenght)

            for i in range(len(column_name_len_treeview)):
                len_treeview[i]=max([len_treeview[i],column_name_len_treeview[i]])            
            
            treeview=ttk.Treeview(self.frame3, columns=column, displaycolumns=column)
    
            scroll_x = ttk.Scrollbar(self.frame3, orient="horizontal", command=treeview.xview)
            scroll_x.pack(side='bottom', fill='x')
            treeview.configure(xscrollcommand=scroll_x.set)

            scroll_y = ttk.Scrollbar(self.frame3, orient="vertical", command=treeview.yview)
            scroll_y.pack(side='right', fill='y')
            treeview.configure(yscrollcommand=scroll_y.set)

            treeview.column("#0", width=40, anchor="center")
            treeview.heading("#0", text="index", anchor="center")

            treeview.pack()

            for i in range(len(column)):
                treeview.column("#%d"%(i+1), width=len_treeview[i], anchor="center")
                treeview.heading(i+1, text=column_name[i], anchor="center")

            for i in range(len((self.selec[self.jud_frame2])[self.jud_frame3])):
                treeview.insert('', 'end', text=i, values=list(self.data[((self.selec[self.jud_frame2])[self.jud_frame3])[i]]), iid=str(i)+"번")

        elif self.jud_frame3==None:
            scrollbar=tk.Scrollbar(self.frame3)
            scrollbar.pack(side="right",fill="y")
            scrollbar_2=tk.Scrollbar(self.frame3,orient=tk.HORIZONTAL)
            scrollbar_2.pack(side="bottom",fill="x")
            listbox=tk.Listbox(self.frame3, width=90,height=20, selectmode="extended", xscrollcommand=scrollbar_2.set,yscrollcommand=scrollbar.set,font=("TkDefaultFont",10))
            listbox.bind('<Double-1>')
            listbox.pack(fill="both",expand=True)
            scrollbar["command"]=listbox.yview
            scrollbar_2["command"]=listbox.xview

    def Frame4(self):
        try :
            self.frame4.destroy()
            self.frame4=tk.Frame(self.bottom_frame, width=150,height=200, relief="sunken", bd=2)
            self.frame4.pack(side="right",fill="both",expand=True)
        except :
            self.frame4=tk.Frame(self.bottom_frame, width=150,height=200, relief="sunken", bd=2)
            self.frame4.pack(side="right",fill="both",expand=True)
        tk.Button(self.frame4, text="정보확인",font=("TkDefaultFont",15,"bold"),command=lambda:self.checking()).pack()
        tk.Button(self.frame4, text="넘어가기",font=("TkDefaultFont",15,"bold"),command=lambda:self.deleting()).pack()
        tk.Button(self.frame4, text="짧게변경",font=("TkDefaultFont",15,"bold"),command=lambda:self.all_setting_short()).pack()
        tk.Button(self.frame4, text="길게변경",font=("TkDefaultFont",15,"bold"),command=lambda:self.all_setting_long()).pack()
        if len(self.master.temp_4)>=1:
            tk.Button(self.frame4, text="다음으로",font=("TkDefaultFont",15,"bold"),command=lambda: self.next()).pack()
        tk.Button(self.frame4, text="EXIT", font=("TkDefaultFont",15,"bold"),command=lambda: self._exit()).pack()

    def _exit(self):
        MsgBox = tk.messagebox.askquestion ('종료','저장하시겠습니까?\n("예" 를 누르시면 현재까지 진행된 내용이 바뀝니다.)')
        if MsgBox=="yes":
            self.save_excel()
            sys.exit()

    def next(self):
        MsgBox = tk.messagebox.askquestion ('다음','저장하시겠습니까?\n("예" 를 누르시면 변경된 내용이 저장된 후 다음 파일로 넘어갑니다.)')
        if MsgBox=="yes":
            self.save_excel()
            self.delete_Frame()
            self.master.switch_frame(PageTwo)
        elif MsgBox=="no":
            self.delete_Frame()
            self.master.switch_frame(PageTwo)

    def save_excel(self):
        for i in list(self.selec.keys()):
            for j in list(self.selec[str(i)].keys()):
                for k in (self.selec[str(i)])[str(j)]:
                    self.data[k,5]=i
        for i in range(len(self.data[:,5])):
            if self.data[i,5] != None:
                k=6+i
                self.all_values[k][5]=self.data[i,5]

        wb=openpyxl.Workbook()
        ws=wb.active
        for i in self.all_values:
            ws.append(i)

        wb.save('엑셀 나오는 곳\\수정 후_'+self.pre.split(".")[0]+".xlsx")

    def checking(self):
        self.jud_frame3=self.listbox2.selection_get()
        self.Frame3()

    def deleting(self):
        if self.jud_frame2!=None:
            MsgBox = tk.messagebox.askquestion ('삭제','정말로 넘어가시겠습니까?\n("예" 를 누르시면 변경 목록에서 삭제됩니다)')
            if MsgBox=="yes":
                self.selec.pop(str(self.jud_frame2))
                self.jud_frame2=None
                self.jud_frame3=None
                self.Frame1()
                self.Frame2()
                self.Frame3()

    def all_setting_short(self):
        MsgBox = tk.messagebox.askquestion ('일괄변경','정말로 일괄변경 하시겠습니까?\n"예" 를 누르시면 가장 짧은 내용이 선택됩니다.')
        if MsgBox=="yes":
            self.jud_frame2=None
            self.jud_frame3=None
            new={}
            for i in self.selec.keys():
                new[str(min(list(self.selec[str(i)].keys()),key=len))]=self.selec[str(i)]
            self.selec=new
            self.Frame1()
            self.Frame2()
            self.Frame3()      

    def all_setting_long(self):
        MsgBox = tk.messagebox.askquestion ('일괄변경','정말로 일괄변경 하시겠습니까?\n"예" 를 누르시면 가장 긴 내용이 선택됩니다.')
        if MsgBox=="yes":
            self.jud_frame2=None
            self.jud_frame3=None
            new={}
            for i in self.selec.keys():
                new[str(max(list(self.selec[str(i)].keys()),key=len))]=self.selec[str(i)]
            self.selec=new
            self.Frame1()
            self.Frame2()
            self.Frame3()

if __name__ == "__main__":
    app = SampleApp()
    app.mainloop()